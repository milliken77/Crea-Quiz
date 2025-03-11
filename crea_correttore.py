import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

def create_exam_correction_sheet(input_file, output_file):
    try:
        # Caricamento del file di input
        with open(input_file, 'r') as file:
            lines = file.readlines()

        # Parsing dei dati di correzione, rimuovendo eventuali spazi
        corrections = []
        for line in lines:
            corrections.append([x.strip() for x in line.strip().split(',')])

        if not corrections:
            print("Errore: Il file di input è vuoto.")
            return

        # Creazione DataFrame
        num_questions = len(corrections[0]) - 1
        correction_df = pd.DataFrame(corrections)
        correction_df.columns = ['Numero Compito'] + [str(i + 1) for i in range(num_questions)]

        # Creazione del workbook Excel
        workbook = Workbook()

        # Tabella Risposte Corrette
        sheet1 = workbook.active
        sheet1.title = "Risposte Corrette"
        for r_idx, row in enumerate(correction_df.values, start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)

        # Formattazione delle colonne per la prima tabella
        for col_idx in range(2, 2 + num_questions):  # Dalla seconda colonna in avanti
            col_letter = get_column_letter(col_idx)
            sheet1.column_dimensions[col_letter].width = 3  # Imposta larghezza a 3 per le risposte

        # Tabella Risposte Studenti
        sheet2 = workbook.create_sheet(title="Risposte Studenti")
        header = ['Numero Compito', 'Nome', 'Cognome'] + [str(i + 1) for i in range(num_questions)] + ['Domande Corrette', 'Voto (%)']
        for c_idx, header_value in enumerate(header, start=1):
            sheet2.cell(row=1, column=c_idx, value=header_value)

        # Popolamento della colonna 'Numero Compito'
        for r_idx in range(2, len(correction_df) + 2):
            sheet2.cell(row=r_idx, column=1, value=int(correction_df.iloc[r_idx - 2, 0]))

        # Formattazione delle colonne
        for col_idx in range(4, 4 + num_questions):
            col_letter = get_column_letter(col_idx)
            sheet2.column_dimensions[col_letter].width = 4

        # Creazione colonne nascoste per verificare le risposte
        hidden_col_start = 4 + num_questions + 2
        for q_idx in range(num_questions):
            col_letter = get_column_letter(hidden_col_start + q_idx)
            sheet2.cell(row=1, column=hidden_col_start + q_idx, value=f"Check {q_idx + 1}")
            sheet2.column_dimensions[col_letter].hidden = True

        # Inserimento formule, gestione calcolo e formattazione condizionale
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        for r_idx in range(2, len(correction_df) + 2):
            for q_idx in range(num_questions):
                # Celle per le risposte degli studenti
                student_answer_cell = sheet2.cell(row=r_idx, column=4 + q_idx)
                student_answer_cell.value = None  # Celle inizialmente vuote per l'input

                # Colonna nascosta per il controllo
                correct_answer = correction_df.iloc[r_idx - 2, q_idx + 1]
                hidden_col = get_column_letter(hidden_col_start + q_idx)
                sheet2[f'{hidden_col}{r_idx}'] = f'=IF({get_column_letter(4 + q_idx)}{r_idx}="{correct_answer}", "X", "")'

                # Formattazione condizionale: risposte corrette in verde
                formula = f'{get_column_letter(4 + q_idx)}{r_idx}="{correct_answer}"'
                sheet2.conditional_formatting.add(
                    f'{get_column_letter(4 + q_idx)}{r_idx}',
                    FormulaRule(formula=[formula], fill=green_fill)
                )

            # Formula per conteggio risposte corrette
            count_start = get_column_letter(hidden_col_start)
            count_end = get_column_letter(hidden_col_start + num_questions - 1)
            sheet2[f'{get_column_letter(4 + num_questions)}{r_idx}'] = f'=COUNTIF({count_start}{r_idx}:{count_end}{r_idx}, "X")'

            # Formula per calcolo del voto
            sheet2[f'{get_column_letter(5 + num_questions)}{r_idx}'] = (
                f'=ROUND(({get_column_letter(4 + num_questions)}{r_idx}/{num_questions})*100, 2)'
            )

        # Salvataggio del file Excel
        workbook.save(output_file)
        print(f"File Excel salvato come '{output_file}'.")

    except Exception as e:
        print(f"Errore durante la creazione del file: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Utilizzo: python correttore.py <input_file> <output_file>")
    else:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
        create_exam_correction_sheet(input_file, output_file)

